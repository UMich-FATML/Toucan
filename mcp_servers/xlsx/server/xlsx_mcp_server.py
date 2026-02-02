"""
XLSX Processing MCP Server

Provides tools for Excel spreadsheet processing, including:
- Recalculating formulas using LibreOffice
- Extracting workbook metadata and sheet information
- Extracting cell values from sheets and ranges
- Validating formula syntax
"""

import sys
from pathlib import Path
from typing import Any

from mcp.server.fastmcp import FastMCP

# Add parent directory to path for imports
PARENT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(PARENT_DIR))

mcp = FastMCP("xlsx-processing")


@mcp.tool(
    annotations={
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": True,
    }
)
def xlsx_recalculate_formulas(xlsx_path: str, timeout: int = 30) -> dict[str, Any]:
    """
    Recalculate all formulas in an Excel file using LibreOffice and detect errors.

    This tool uses LibreOffice to recalculate all formulas in the workbook and
    then scans for Excel error values (#VALUE!, #DIV/0!, #REF!, etc.).

    Args:
        xlsx_path: Path to the Excel file to recalculate
        timeout: Maximum time to wait for recalculation in seconds (default: 30)

    Returns:
        A dictionary with:
        - success: Whether the operation completed successfully
        - status: 'success' if no errors, 'errors_found' if errors detected
        - total_errors: Total count of Excel errors found
        - total_formulas: Number of formulas in the workbook
        - error_summary: Breakdown by error type with locations
        - message: Human-readable description of the result
    """
    try:
        from recalc import recalc

        result = recalc(xlsx_path, timeout)

        if 'error' in result:
            return {
                "success": False,
                "status": "error",
                "total_errors": 0,
                "total_formulas": 0,
                "error_summary": {},
                "message": result['error'],
            }

        return {
            "success": True,
            "status": result.get('status', 'unknown'),
            "total_errors": result.get('total_errors', 0),
            "total_formulas": result.get('total_formulas', 0),
            "error_summary": result.get('error_summary', {}),
            "message": f"Recalculation complete: {result.get('total_formulas', 0)} formulas processed, {result.get('total_errors', 0)} errors found",
        }
    except Exception as e:
        return {
            "success": False,
            "status": "error",
            "total_errors": 0,
            "total_formulas": 0,
            "error_summary": {},
            "message": f"Error recalculating formulas: {str(e)}",
        }


@mcp.tool(
    annotations={
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    }
)
def xlsx_get_sheet_info(xlsx_path: str) -> dict[str, Any]:
    """
    Get workbook metadata including sheet names, dimensions, and formula counts.

    This tool extracts structural information about an Excel workbook without
    modifying it, useful for understanding the workbook layout before processing.

    Args:
        xlsx_path: Path to the Excel file to analyze

    Returns:
        A dictionary with:
        - success: Whether the operation completed successfully
        - sheet_count: Number of sheets in the workbook
        - sheets: List of sheet info (name, dimensions, formula_count)
        - total_formulas: Total formulas across all sheets
        - message: Human-readable description of the result
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, data_only=False)

        sheets = []
        total_formulas = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            formula_count = 0

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1

            total_formulas += formula_count

            sheets.append({
                "name": sheet_name,
                "dimensions": ws.dimensions if ws.dimensions else "A1:A1",
                "formula_count": formula_count,
            })

        wb.close()

        return {
            "success": True,
            "sheet_count": len(sheets),
            "sheets": sheets,
            "total_formulas": total_formulas,
            "message": f"Workbook has {len(sheets)} sheet(s) with {total_formulas} total formulas",
        }
    except Exception as e:
        return {
            "success": False,
            "sheet_count": 0,
            "sheets": [],
            "total_formulas": 0,
            "message": f"Error reading workbook: {str(e)}",
        }


@mcp.tool(
    annotations={
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    }
)
def xlsx_extract_cell_values(
    xlsx_path: str,
    sheet_name: str = None,
    cell_range: str = None,
    data_only: bool = True
) -> dict[str, Any]:
    """
    Extract cell values from a sheet or specific range in an Excel file.

    This tool reads cell values from the workbook. Use data_only=True to get
    calculated values (after formulas are computed), or data_only=False to
    see the raw formulas.

    Args:
        xlsx_path: Path to the Excel file to read
        sheet_name: Name of sheet to read (default: first sheet)
        cell_range: Cell range to extract, e.g., "A1:D10" (default: all used cells)
        data_only: If True, return calculated values; if False, return formulas

    Returns:
        A dictionary with:
        - success: Whether the operation completed successfully
        - data: 2D array of cell values
        - dimensions: The range of cells extracted
        - sheet_name: Name of the sheet read
        - message: Human-readable description of the result
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import range_boundaries

        wb = load_workbook(xlsx_path, data_only=data_only)

        # Select sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return {
                    "success": False,
                    "data": [],
                    "dimensions": "",
                    "sheet_name": sheet_name,
                    "message": f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}",
                }
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        # Determine range to extract
        if cell_range:
            try:
                min_col, min_row, max_col, max_row = range_boundaries(cell_range)
            except Exception:
                wb.close()
                return {
                    "success": False,
                    "data": [],
                    "dimensions": "",
                    "sheet_name": sheet_name,
                    "message": f"Invalid cell range: {cell_range}",
                }
            dimensions = cell_range
        else:
            min_row, min_col = 1, 1
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            dimensions = ws.dimensions if ws.dimensions else "A1:A1"

        # Extract data
        data = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            row_data = []
            for cell in row:
                value = cell.value
                # Convert to JSON-serializable types
                if value is None:
                    row_data.append(None)
                elif isinstance(value, (int, float, str, bool)):
                    row_data.append(value)
                else:
                    row_data.append(str(value))
            data.append(row_data)

        wb.close()

        return {
            "success": True,
            "data": data,
            "dimensions": dimensions,
            "sheet_name": sheet_name,
            "message": f"Extracted {len(data)} rows from sheet '{sheet_name}'",
        }
    except Exception as e:
        return {
            "success": False,
            "data": [],
            "dimensions": "",
            "sheet_name": sheet_name or "",
            "message": f"Error extracting cell values: {str(e)}",
        }


@mcp.tool(
    annotations={
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    }
)
def xlsx_validate_formulas(xlsx_path: str) -> dict[str, Any]:
    """
    Validate formula syntax in an Excel file without recalculating.

    This tool checks formulas for common issues like existing error values,
    empty formulas, and basic syntax problems. It does not use LibreOffice
    and does not recalculate formulas.

    Args:
        xlsx_path: Path to the Excel file to validate

    Returns:
        A dictionary with:
        - success: Whether the operation completed successfully
        - is_valid: True if no issues found, False otherwise
        - issues: List of validation issues found
        - formula_count: Total number of formulas checked
        - message: Human-readable description of the result
    """
    try:
        from openpyxl import load_workbook

        # Load with data_only=False to see formulas
        wb_formulas = load_workbook(xlsx_path, data_only=False)
        # Load with data_only=True to see calculated values (for error detection)
        wb_values = load_workbook(xlsx_path, data_only=True)

        issues = []
        formula_count = 0
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']

        for sheet_name in wb_formulas.sheetnames:
            ws_formulas = wb_formulas[sheet_name]
            ws_values = wb_values[sheet_name]

            for row_idx, row in enumerate(ws_formulas.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    cell_value = cell.value

                    # Check if it's a formula
                    if cell_value and isinstance(cell_value, str) and cell_value.startswith('='):
                        formula_count += 1
                        location = f"{sheet_name}!{cell.coordinate}"

                        # Check for empty formula (just "=")
                        if cell_value.strip() == '=':
                            issues.append({
                                "type": "empty_formula",
                                "location": location,
                                "description": "Empty formula (just '=')",
                            })

                        # Check for unbalanced parentheses
                        if cell_value.count('(') != cell_value.count(')'):
                            issues.append({
                                "type": "unbalanced_parentheses",
                                "location": location,
                                "description": "Unbalanced parentheses in formula",
                            })

                    # Check calculated values for errors
                    try:
                        value_cell = ws_values.cell(row=row_idx, column=col_idx)
                        calc_value = value_cell.value
                        if calc_value and isinstance(calc_value, str):
                            for err in excel_errors:
                                if err in calc_value:
                                    location = f"{sheet_name}!{cell.coordinate}"
                                    issues.append({
                                        "type": "error_value",
                                        "location": location,
                                        "description": f"Cell contains error: {err}",
                                        "error_type": err,
                                    })
                                    break
                    except Exception:
                        pass

        wb_formulas.close()
        wb_values.close()

        is_valid = len(issues) == 0

        return {
            "success": True,
            "is_valid": is_valid,
            "issues": issues[:50],  # Limit to first 50 issues
            "formula_count": formula_count,
            "message": f"Validated {formula_count} formulas, found {len(issues)} issue(s)",
        }
    except Exception as e:
        return {
            "success": False,
            "is_valid": False,
            "issues": [],
            "formula_count": 0,
            "message": f"Error validating formulas: {str(e)}",
        }


if __name__ == "__main__":
    mcp.run()
